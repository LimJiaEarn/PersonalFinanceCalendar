import json, datetime, calendar, openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, numbers
from openpyxl.formatting.rule import Rule, FormulaRule, CellIsRule

# C:\Users\user\AppData\Local\Programs\Python\Python311\Scripts

#Utility lists
daysOfWeek = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
INPUT_DICT = json.load(open("YourCalendarSettings.txt"))
YEAR = INPUT_DICT[list(INPUT_DICT)[0]]
startDayOfMonth = calendar.day_name[datetime.datetime.strptime('01 01 '+str(YEAR), '%d %m %Y').weekday() ]
startDayOfMonth_index = daysOfWeek.index(startDayOfMonth) #initially set to start day of jan(start of new year), re-used for subsequent months
#Index 0 - YEAR
#Index 1 - CalendarDaysAndMonths
#Index 2 - START_DAY -> 0 is Monday, 6 is Sunday
#Index 3 - left_headerTitles
#Index 4 - right_headerTitles
#Index 5 - week titles
#Index 6 - Main Inflow Overview
#Index 7 - Main Expenses Overview 
#Index 8 - Inflow Breakdown
#Index 9 - Outflow Breakdown
#Index 10 - Theme Colors

#Rows needed to offset before the first row of main data, absolute +3 because of seperator cells
rowOffset_header = max(len(INPUT_DICT[list(INPUT_DICT)[3]]), len(INPUT_DICT[list(INPUT_DICT)[4]])) +3
#Rows needed to offset in each week's block, absolute +1 because of seperator cells
rowOffset_weeks = len(INPUT_DICT[list(INPUT_DICT)[5]]) +1

#Columns set as default 2 to include some aesthetics
columnOffset = 2

#Pre-set cell styling and colors (not user changing yet through code)
summaryExpensesSideFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[0])
summaryTotalRowFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[1])
summaryDepthRowFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[2])
seperatorFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[3])
sheetMonthFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[4])
sheetDaysFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[5])
headerSideFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[6])
leftheaderTitleFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[7])
rightheaderTitleFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[8])
weekSideFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[9])
weekTitleFill = [PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[10][0]), PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[10][1]),\
                PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[10][2]), PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[10][3]),\
                PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[10][4]), PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[10][5])]
DateFill = PatternFill(fill_type="solid", start_color=list(INPUT_DICT[list(INPUT_DICT)[10]])[11])

            
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

#Loads a user excel file, if unable to find, creates a new empty excel file
FileName = input("Enter your excel file name: ")
try:
    CalendarFile = openpyxl.load_workbook(FileName+".xlsx")
except:
    CalendarFile = openpyxl.Workbook() 



#Delete possible default sheets created
if ("Sheet1" in CalendarFile.sheetnames):
    del CalendarFile["Sheet1"]
if ("Sheet" in CalendarFile.sheetnames):
    del CalendarFile["Sheet"]

#Utility function to check if any value exist in row
def checkRow(row):
    for i in range(2, 9):
        if (CalendarFile_currentSheet[get_column_letter(columnOffset+i)+str(row)].value):
            return True
    return False

#Generate Overview Page
#Further initialisations are done after month sheet intialisation
CalendarFile.create_sheet(str(YEAR)+" Overview")

#Traversing through every month and generate its content
for month in INPUT_DICT[list(INPUT_DICT)[1]].keys():
    
    #Create a sheet for each month
    CalendarFile.create_sheet(month[0:3])
    CalendarFile_currentSheet = CalendarFile[month[0:3]]
    
    #Generate styling for Header
    #Absolute referencing is used for simplicity as certain cells will not be affected by user preference
    cell = get_column_letter(columnOffset)+"1"
    CalendarFile_currentSheet.merge_cells(cell+":J1")
    CalendarFile_currentSheet[cell] = month.upper() + "  "+ str(YEAR)
    CalendarFile_currentSheet[cell].font = Font(bold=True, size = 22)
    CalendarFile_currentSheet[cell].alignment = Alignment(horizontal='center', vertical='center')
    CalendarFile_currentSheet[cell].fill = sheetMonthFill
    CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+"2:J2")
    CalendarFile_currentSheet[get_column_letter(columnOffset)+"2"].fill = seperatorFill
    CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(rowOffset_header)+":"+"J"+str(rowOffset_header))
    CalendarFile_currentSheet[get_column_letter(columnOffset)+str(rowOffset_header)].fill = seperatorFill
    
    #Generating left header titles
    row=3
    for title, accounting in INPUT_DICT[list(INPUT_DICT)[3]].items():
        CalendarFile_currentSheet[get_column_letter(columnOffset+1)+str(row)] = title
        CalendarFile_currentSheet[get_column_letter(columnOffset+1)+str(row)].fill = leftheaderTitleFill
        for i in range(3):
            CalendarFile_currentSheet[get_column_letter(columnOffset+i)+str(row)].border = thin_border
        CalendarFile_currentSheet[get_column_letter(columnOffset+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
        #This prepares the accounting row
        if (accounting==1 or accounting==2):
            CalendarFile_currentSheet[get_column_letter(columnOffset+2)+str(row)]=1 #Change back to 0 eventually
        row+=1

    #Generating right header titles
    row=3
    for title, accounting in INPUT_DICT[list(INPUT_DICT)[4]].items():
        CalendarFile_currentSheet[get_column_letter(columnOffset+5)+str(row)]=title
        CalendarFile_currentSheet[get_column_letter(columnOffset+5)+str(row)].fill = rightheaderTitleFill
        for i in range(2):
            CalendarFile_currentSheet[get_column_letter(columnOffset+5+i)+str(row)].border = thin_border
        CalendarFile_currentSheet[get_column_letter(columnOffset+5)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
        #This prepares the accounting row
        if (accounting==1 or accounting==2):
            CalendarFile_currentSheet[get_column_letter(columnOffset+6)+str(row)]=1 #Change back to 0 eventually
        row+=1

    CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+"3"+":"+get_column_letter(columnOffset)+str(rowOffset_header-1))
    CalendarFile_currentSheet[get_column_letter(columnOffset)+"3"].fill = headerSideFill
    CalendarFile_currentSheet.column_dimensions[get_column_letter(columnOffset)].width = 2

    #Generating days of each week based on user preset preference
    row = rowOffset_header + 1
    start_day = INPUT_DICT[list(INPUT_DICT)[2]]
    CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(row)+":"+get_column_letter(columnOffset+1)+str(row))
    CalendarFile_currentSheet[get_column_letter(columnOffset)+str(row)].fill = sheetDaysFill
    for column in range(2, 9):
        char = get_column_letter(columnOffset+column)
        CalendarFile_currentSheet[char+str(row)] = daysOfWeek[start_day]
        CalendarFile_currentSheet[char+str(row)].alignment = Alignment(horizontal='center', vertical='center')
        CalendarFile_currentSheet[char+str(row)].fill = sheetDaysFill
        CalendarFile_currentSheet[char+str(row)].font = Font(bold=True, size=11)
        start_day = (start_day+1)%7
    
    #Seperator
    CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(row+1)+":"+get_column_letter(columnOffset+8)+str(row+1))
    CalendarFile_currentSheet[get_column_letter(columnOffset)+str(row+1)].fill = seperatorFill
    CalendarFile_currentSheet.row_dimensions[row+1].height = 20

    #Generate dates of each week with respective to its day as set above
    date=1
    for row in range(rowOffset_header+3, rowOffset_header + rowOffset_weeks*6, rowOffset_weeks):
        
        for column in range(2, 9):              
            char = get_column_letter(columnOffset+column)
            
            #Ensure first date of month is inserted in the correct starting day
            if ( (date==1) and (CalendarFile_currentSheet[char+str(rowOffset_header+1)].value != startDayOfMonth) ):
                continue
            #Continuos insertion
            elif (date<=INPUT_DICT[list(INPUT_DICT)[1]][month]):   
                CalendarFile_currentSheet[char+str(row)] = date
                CalendarFile_currentSheet[char+str(row)].font = Font(bold=True)
                CalendarFile_currentSheet[char+str(row)].fill = DateFill
                CalendarFile_currentSheet[char+str(row)].alignment = Alignment(horizontal='center', vertical='center')
                date+=1
            #When all dates of the month has been filled
            elif (date==INPUT_DICT[list(INPUT_DICT)[1]][month]+1): 
                startDayOfMonth = CalendarFile_currentSheet[char+str(rowOffset_header+1)].value
                date+=1 
                break
            #Simply break through remaining loops
            else:
                break

    #Generate week title for each week
    row = rowOffset_header+3
    weekSideFillIndex = 0
    while(checkRow(row)):
        #Aesthetic bookmarklike column
        CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(row)+":"+get_column_letter(columnOffset)+str(row+len(INPUT_DICT[list(INPUT_DICT)[5]])-1))
        CalendarFile_currentSheet[get_column_letter(columnOffset)+str(row)].fill = weekSideFill
        CalendarFile_currentSheet.column_dimensions[get_column_letter(columnOffset)].width = 2
        
        #Slotting titles
        for title in INPUT_DICT[list(INPUT_DICT)[5]].keys():
            CalendarFile_currentSheet[get_column_letter(columnOffset+1)+str(row)]=title
            CalendarFile_currentSheet.row_dimensions[row].height = 18
            CalendarFile_currentSheet[get_column_letter(columnOffset+1)+str(row)].fill=weekTitleFill[weekSideFillIndex%6]
            for k in range(8):
                CalendarFile_currentSheet[get_column_letter(columnOffset+1+k)+str(row)].border = thin_border
            CalendarFile_currentSheet[get_column_letter(columnOffset+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
            #This prepares the accounting row
            if (INPUT_DICT[list(INPUT_DICT)[5]][title]==1):
                for i in range(2, 9):
                    CalendarFile_currentSheet[get_column_letter(columnOffset+i)+str(row)] = 1 #change back to 0 eventually
            if (INPUT_DICT[list(INPUT_DICT)[5]][title]==2):
                for i in range(2, 9):
                    CalendarFile_currentSheet[get_column_letter(columnOffset+i)+str(row)] = 1 #change back to 0 eventually
            
            row+=1
        CalendarFile_currentSheet.row_dimensions[row-len(INPUT_DICT[list(INPUT_DICT)[5]])].height = 28
        CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(row)+":"+get_column_letter(columnOffset+8)+str(row))
        CalendarFile_currentSheet[get_column_letter(columnOffset)+str(row)].fill = seperatorFill
        row+=1
        weekSideFillIndex+=1

    #Setting height for header portion
    CalendarFile_currentSheet.row_dimensions[1].height = 30
    for i in range(rowOffset_header):
        CalendarFile_currentSheet.row_dimensions[3+i].height = 20
        
    #Setting height for weeks layout
    row = rowOffset_header + 3
    while (CalendarFile_currentSheet[get_column_letter(columnOffset+1)+str(row)].value==INPUT_DICT[list(INPUT_DICT)[5]][list(INPUT_DICT[list(INPUT_DICT)[5]])[0]]): 
        for i in range(rowOffset_weeks):
            CalendarFile_currentSheet.row_dimensions[row+i].height = 35
        row+=rowOffset_weeks
    
    #Setting width for side column 
    CalendarFile_currentSheet.column_dimensions[get_column_letter(columnOffset-1)].width = 8
    CalendarFile_currentSheet.column_dimensions[get_column_letter(columnOffset+1)].width = 20
    
    #Setting width for days of week
    for i in range(2, 9):
        CalendarFile_currentSheet.column_dimensions[get_column_letter(columnOffset+i)].width = 25
    CalendarFile_currentSheet.row_dimensions[rowOffset_header + 1].height = 20


def findAccCellinHeader(title):
    #Left header titles
    if (title in INPUT_DICT[list(INPUT_DICT)[3]]): 
        return get_column_letter(1+columnOffset+1)+str(3+list(INPUT_DICT[list(INPUT_DICT)[3]]).index(title))
    #Right header titles
    elif (title in INPUT_DICT[list(INPUT_DICT)[4]]): 
        return get_column_letter(1+columnOffset+5)+str(3+list(INPUT_DICT[list(INPUT_DICT)[4]]).index(title))
    
    return -1
    

#Returns first row in occurs as an integer
def findAccCellinWeek(title):
    row=rowOffset_header+3
    if (title in INPUT_DICT[list(INPUT_DICT)[5]]): 
        # get_column_letter(columnOffset+1)+str(row)
        return row+list(INPUT_DICT[list(INPUT_DICT)[5]]).index(title)
    
    return -1


#OVERVIEW


rowOffset_Overview = 2
rowOffset_OverviewBody = 2+len(INPUT_DICT[list(INPUT_DICT)[6]])+len(INPUT_DICT[list(INPUT_DICT)[7]])+2+2 #2 for total, 2 for spacing
#Columns set as default 2 to include some aesthetics
columnOffset_Overview = 2

CalendarFile_currentSheet = CalendarFile[str(YEAR)+" Overview"]
CalendarFile_currentSheet.merge_cells("B1:F1")
CalendarFile_currentSheet.merge_cells("G1:O1")
CalendarFile_currentSheet["B1"] = str(YEAR)+" Overview"
for i in range(2, 16):
    CalendarFile_currentSheet[get_column_letter(i)+"1"].border = Border(bottom=Side(style='thick'))
CalendarFile_currentSheet["B1"].font = Font(bold=True, size = 22)
CalendarFile_currentSheet["B1"].alignment = Alignment(horizontal='center', vertical='center')
CalendarFile_currentSheet["B1"].fill = sheetMonthFill
CalendarFile_currentSheet.column_dimensions[get_column_letter(columnOffset_Overview+1)].width = 25
CalendarFile_currentSheet.column_dimensions[get_column_letter(columnOffset_Overview)].width = 2
CalendarFile_currentSheet.column_dimensions[get_column_letter(1)].width = 5

#Generating months
m=1
row=rowOffset_Overview+2
CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(rowOffset_Overview)+":"+get_column_letter(columnOffset+13)+str(rowOffset_Overview))
for month in INPUT_DICT[list(INPUT_DICT)[1]].keys():
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+m)+str(row-1)] = month[0:3]
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+m)+str(row-1)].alignment = Alignment(horizontal='right', vertical='center')
    m+=1

#Overview Inflow
CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset_Overview)+str(row-1)+":"+get_column_letter(columnOffset_Overview)+str(row+len(INPUT_DICT[list(INPUT_DICT)[6]])))
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview)+str(row-1)].fill = leftheaderTitleFill
for tup in INPUT_DICT[list(INPUT_DICT)[6]]:
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)] = tup[0]
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
    #For non summation rows
    if (len(tup[3])==0):

        if (tup[2]==1): #Header type
            cell = findAccCellinHeader(tup[0])
            if (cell!=-1):
                for i in range(12):
                    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = "=SUM("+str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])+'!'+str(cell)+')'

        elif (tup[2]==2): #Week type
            for i in range(12):
                cellRow = findAccCellinWeek(tup[0])
                if (cellRow!=-1):
                    cells="=SUM("
                    while True:
                        month = str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])
                        cells+=(month+'!'+get_column_letter(columnOffset+2)+str(cellRow)+':'+get_column_letter(columnOffset+8)+str(cellRow)+',')
                        cellRow+=rowOffset_weeks
                        if (CalendarFile[month][get_column_letter(columnOffset+1)+str(cellRow)].value!=tup[0]):
                            break
                    cells=cells[:-1]+')'
                    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = cells   
    #For summation rows
    else:
        #For every month
        for i in range(12):
            sumCode="="
            month = str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])
            #For every title inside summation row
            for titleTup in tup[3]:

                #Header title
                if (titleTup[1] == 1):
                                            
                    cell = findAccCellinHeader(titleTup[0])
                    if (cell!=-1):
                        #Iterating through each month
                        sumCode2 = "SUM("+month+'!'+str(cell)+')'
                        sumCode+=(sumCode2+"+")
                
                #Week title
                elif (titleTup[1] == 2):
                    
                        cellRow = findAccCellinWeek(titleTup[0])
                        if (cellRow!=-1):
                            sumCode2="SUM("
                            while (CalendarFile[month][get_column_letter(columnOffset+1)+str(cellRow)].value==titleTup[0]):
                                sumCode2+=(month+'!'+get_column_letter(columnOffset+2)+str(cellRow)+':'+get_column_letter(columnOffset+8)+str(cellRow)+',')
                                cellRow+=rowOffset_weeks

                            sumCode2=sumCode2[:-1]+')'
                            sumCode+=(sumCode2+"+")
                            
            sumCode=sumCode[:-1]
            CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = sumCode  



    row+=1

if (len(INPUT_DICT[list(INPUT_DICT)[6]])!=0):
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)] = "Total"
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].fill = summaryTotalRowFill
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
    for m2 in range(12):
        column = get_column_letter(columnOffset_Overview+2+m2)
        CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)] = "=SUM("+column+str(row-len(INPUT_DICT[list(INPUT_DICT)[6]]))+":"+column+str(row-1)+")"
        CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)].fill = summaryTotalRowFill
        CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)].alignment = Alignment(horizontal='right', vertical='center')
CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(row+1)+":"+get_column_letter(columnOffset+13)+str(row+1))

#Overview Outflow
row+=2
CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset_Overview)+str(row)+":"+get_column_letter(columnOffset_Overview)+str(row+len(INPUT_DICT[list(INPUT_DICT)[7]])))
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview)+str(row)].fill = summaryExpensesSideFill
for tup in INPUT_DICT[list(INPUT_DICT)[7]]:
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)] = tup[0]
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
    if (len(tup[3])==0):
        if (tup[2]==1): #Header type
            cell = findAccCellinHeader(tup[0])
            if (cell!=-1):
                for i in range(12):
                    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = "=SUM("+str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])+'!'+str(cell)+')'

        elif (tup[2]==2): #Week type
            for i in range(12):
                cellRow = findAccCellinWeek(tup[0])
                if (cellRow!=-1):
                    cells="=SUM("
                    while True:
                        month = str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])
                        cells+=(month+'!'+get_column_letter(columnOffset+2)+str(cellRow)+':'+get_column_letter(columnOffset+8)+str(cellRow)+',')
                        cellRow+=rowOffset_weeks
                        if (CalendarFile[month][get_column_letter(columnOffset+1)+str(cellRow)].value!=tup[0]):
                            break
                    cells=cells[:-1]+')'
                    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = cells   
    
    #For summation rows
    else:
        #For every month
        for i in range(12):
            sumCode="="
            month = str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])
            #For every title inside summation row
            for titleTup in tup[3]:

                #Header title
                if (titleTup[1] == 1):
                                            
                    cell = findAccCellinHeader(titleTup[0])
                    if (cell!=-1):
                        #Iterating through each month
                        sumCode2 = "SUM("+month+'!'+str(cell)+')'
                        sumCode+=(sumCode2+"+")
                
                #Week title
                elif (titleTup[1] == 2):
                    
                        cellRow = findAccCellinWeek(titleTup[0])
                        if (cellRow!=-1):
                            sumCode2="SUM("
                            while (CalendarFile[month][get_column_letter(columnOffset+1)+str(cellRow)].value==titleTup[0]):
                                sumCode2+=(month+'!'+get_column_letter(columnOffset+2)+str(cellRow)+':'+get_column_letter(columnOffset+8)+str(cellRow)+',')
                                cellRow+=rowOffset_weeks

                            sumCode2=sumCode2[:-1]+')'
                            sumCode+=(sumCode2+"+")
                            
            sumCode=sumCode[:-1]
            CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = sumCode  

    
    row+=1

if (len(INPUT_DICT[list(INPUT_DICT)[7]])):
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)] = "Total"
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].fill = summaryTotalRowFill
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
    for m2 in range(12):
        column = get_column_letter(columnOffset_Overview+2+m2)
        CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)] = "=SUM("+column+str(row-len(INPUT_DICT[list(INPUT_DICT)[7]]))+":"+column+str(row-1)+")"
        CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)].fill = summaryTotalRowFill
        CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)].alignment = Alignment(horizontal='right', vertical='center')

CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(row+1)+":"+get_column_letter(columnOffset+13)+str(row+2))


#SUMMARY

#Rows needed to offset before the first row of general summary
rowOffset_Summary = 2 + rowOffset_OverviewBody
rowOffset_SummaryBody = rowOffset_Summary + len(INPUT_DICT[list(INPUT_DICT)[6]]) + 1 + len(INPUT_DICT[list(INPUT_DICT)[7]])

#Generate first summary body
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(rowOffset_Summary+1)] = "Inflow"
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(rowOffset_Summary+1)].alignment = Alignment(horizontal='center', vertical='center')

#Generating months
row=rowOffset_Summary+2
m=1
for month in INPUT_DICT[list(INPUT_DICT)[1]].keys():
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+m)+str(rowOffset_Summary+1)] = month[0:3]
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+m)+str(rowOffset_Summary+1)].alignment = Alignment(horizontal='right', vertical='center')
    m+=1
for tup in INPUT_DICT[list(INPUT_DICT)[8]]:
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)] = tup[0]
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')

    #For non summation rows
    if (len(tup[3])==0): 
        if (tup[2]==1): #Header type
            cell = findAccCellinHeader(tup[0])
            if (cell!=-1):
                for i in range(12):
                    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = "=SUM("+str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])+'!'+str(cell)+')'

        elif (tup[2]==2): #Week type
            for i in range(12):
                cellRow = findAccCellinWeek(tup[0])
                if (cellRow!=-1):
                    cells="=SUM("
                    while True:
                        month = str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])
                        cells+=(month+'!'+get_column_letter(columnOffset+2)+str(cellRow)+':'+get_column_letter(columnOffset+8)+str(cellRow)+',')
                        cellRow+=rowOffset_weeks
                        if (CalendarFile[month][get_column_letter(columnOffset+1)+str(cellRow)].value!=tup[0]):
                            break

                    cells=cells[:-1]+')'

                    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = cells
    #For summation rows
    else:
        #For every month
        for i in range(12):
            sumCode="="
            month = str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])
            #For every title inside summation row
            for titleTup in tup[3]:

                #Header title
                if (titleTup[1] == 1):
                                            
                    cell = findAccCellinHeader(titleTup[0])
                    if (cell!=-1):
                        #Iterating through each month
                        sumCode2 = "SUM("+month+'!'+str(cell)+')'
                        sumCode+=(sumCode2+"+")
                
                #Week title
                elif (titleTup[1] == 2):
                    
                        cellRow = findAccCellinWeek(titleTup[0])
                        if (cellRow!=-1):
                            sumCode2="SUM("
                            while (CalendarFile[month][get_column_letter(columnOffset+1)+str(cellRow)].value==titleTup[0]):
                                sumCode2+=(month+'!'+get_column_letter(columnOffset+2)+str(cellRow)+':'+get_column_letter(columnOffset+8)+str(cellRow)+',')
                                cellRow+=rowOffset_weeks

                            sumCode2=sumCode2[:-1]+')'
                            sumCode+=(sumCode2+"+")
                            
            sumCode=sumCode[:-1]
            CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = sumCode  

    row+=1
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)] = "Total"
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].fill = summaryTotalRowFill
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset_Overview)+str(rowOffset_Summary+1)+":"+get_column_letter(columnOffset_Overview)+str(row))
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview)+str(rowOffset_Summary+1)].fill = summaryDepthRowFill
#Generating Total row summations
for m2 in range(12):
    column = get_column_letter(columnOffset_Overview+2+m2)
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)] = "=SUM("+column+str(row-len(INPUT_DICT[list(INPUT_DICT)[8]]))+":"+column+str(row-1)+")"
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)].fill = summaryTotalRowFill
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)].alignment = Alignment(horizontal='right', vertical='center')
CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(row+1)+":"+get_column_letter(columnOffset+13)+str(row+1))
#Generate second summary body
row=rowOffset_Summary+2+len(INPUT_DICT[list(INPUT_DICT)[8]])+2
m=1
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)] = "Expenses"
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
#Generating months
for month in INPUT_DICT[list(INPUT_DICT)[1]].keys():
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+m)+str(row)] = month[0:3]
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+m)+str(row)].alignment = Alignment(horizontal='right', vertical='center')
    m+=1
row+=1
for tup in INPUT_DICT[list(INPUT_DICT)[9]]:
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)] = tup[0]
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
    if (len(tup[3])==0):

        if (tup[2]==1): #Header type
            cell = findAccCellinHeader(tup[0])
            if (cell!=-1):
                for i in range(12):
                    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = "=SUM("+str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])+'!'+str(cell)+')'
        elif (tup[2]==2): #Week type
            for i in range(12):
                cellRow = findAccCellinWeek(tup[0])
                if (cellRow!=-1):
                    cells="=SUM("
                    while True:
                        month = str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])
                        cells+=(month+'!'+get_column_letter(columnOffset+2)+str(cellRow)+':'+get_column_letter(columnOffset+8)+str(cellRow)+',')
                        cellRow+=rowOffset_weeks
                        if (CalendarFile[month][get_column_letter(columnOffset+1)+str(cellRow)].value!=tup[0]):
                            break

                    cells=cells[:-1]+')'
                    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = cells

    #For summation rows
    else:
        #For every month
        for i in range(12):
            sumCode="="
            month = str(list(INPUT_DICT[list(INPUT_DICT)[1]].keys())[i][0:3])
            #For every title inside summation row
            for titleTup in tup[3]:

                #Header title
                if (titleTup[1] == 1):
                                            
                    cell = findAccCellinHeader(titleTup[0])
                    if (cell!=-1):
                        #Iterating through each month
                        sumCode2 = "SUM("+month+'!'+str(cell)+')'
                        sumCode+=(sumCode2+"+")
                
                #Week title
                elif (titleTup[1] == 2):
                    
                        cellRow = findAccCellinWeek(titleTup[0])
                        if (cellRow!=-1):
                            sumCode2="SUM("
                            while (CalendarFile[month][get_column_letter(columnOffset+1)+str(cellRow)].value==titleTup[0]):
                                sumCode2+=(month+'!'+get_column_letter(columnOffset+2)+str(cellRow)+':'+get_column_letter(columnOffset+8)+str(cellRow)+',')
                                cellRow+=rowOffset_weeks

                            sumCode2=sumCode2[:-1]+')'
                            sumCode+=(sumCode2+"+")
                            
            sumCode=sumCode[:-1]
            CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1+i+1)+str(row)] = sumCode  
    
    row+=1
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)] = "Total"
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].fill = summaryTotalRowFill
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+1)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset_Overview)+str(row-len(INPUT_DICT[list(INPUT_DICT)[9]])-1)+":"+get_column_letter(columnOffset_Overview)+str(row))
CalendarFile_currentSheet[get_column_letter(columnOffset_Overview)+str(row-len(INPUT_DICT[list(INPUT_DICT)[9]])-1)].fill = summaryDepthRowFill
#Generating Total row summations
for m2 in range(12):
    column = get_column_letter(columnOffset_Overview+2+m2)
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)] = "=SUM("+column+str(row-len(INPUT_DICT[list(INPUT_DICT)[9]]))+":"+column+str(row-1)+")"
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)].fill = summaryTotalRowFill
    CalendarFile_currentSheet[get_column_letter(columnOffset_Overview+2+m2)+str(row)].alignment = Alignment(horizontal='right', vertical='center')
CalendarFile_currentSheet.merge_cells(get_column_letter(columnOffset)+str(row+1)+":"+get_column_letter(columnOffset+13)+str(row+1))

#Save and Exit script
CalendarFile.save(FileName+".xlsx")
print(f"Calendar successfully created with filename: {FileName}!")